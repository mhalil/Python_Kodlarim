"""
Erdal2428 (Erdal ÖZDEMİR) tarafından sorulan sorunun, Erdal bey tarafından düzenlenen kodun son hali;
Link: https://forum.yazbel.com/t/kurum-bilgilerini-guncelleme/21579/10
"""

import pandas as pd
from openpyxl import load_workbook
import datetime
import time

df_ana = pd.read_excel("Ana_Tablo.xlsx")
df_guncel = pd.read_excel("Güncel_Kurum_Sayıları.xlsx")


df_ana["Kurum_Kod_Kod1"] = df_ana["KURUM_KODU"].astype(str) + df_ana["KURUM_KODU1"].astype(str)
df_ana = df_ana.set_index("Kurum_Kod_Kod1")

df_guncel["Kurum_Kod_Kod1"] = df_guncel["KURUM_KODU"].astype(str) + df_guncel["KURUM_KODU1"].astype(str)
df_guncel = df_guncel.set_index("Kurum_Kod_Kod1")

df_ana_indeks = df_ana.index
df_guncel_indeks = df_guncel.index

### Güncelleme Fonksiyonu
def guncelle():
    for indeks in df_ana.index:
        if indeks in df_guncel.index:
            if (df_ana.loc[indeks, "KURUM_ADI"] == df_guncel.loc[indeks, "KURUM_ADI"]):
                continue
            else:
                df_ana.loc[indeks, "KURUM_ADI"] = df_guncel.loc[indeks, "KURUM_ADI"]
                df_ana.loc[indeks, "DURUMU"] = datetime.date.today().__format__('Kurum Adı Değişikliği' +" "+"("+ ('%d.%m.%Y'))+')'

                if (df_ana.loc[indeks, "KURUM_TÜRÜ"]!= df_guncel.loc[indeks, "KURUM_TÜRÜ"]):
                               df_ana.loc[indeks, "KURUM_TÜRÜ"] = df_guncel.loc[indeks, "KURUM_TÜRÜ"]                             
guncelle()

eklenecekler = []
for indeks in df_guncel_indeks:
    if indeks not in df_ana_indeks:
        eklenecekler.append(indeks)

sil = []
for indeks in df_ana_indeks:
    if indeks not in df_guncel_indeks:
        sil.append(indeks)

def satir_sil(satir):
    df_ana.drop(satir, axis = 0, inplace = True)

for satir in sil:
    satir_sil(satir)

def ekle():
    for veri in eklenecekler:
        df_ana.loc[veri] = df_guncel.loc[veri]
        df_ana.loc[veri, "DURUMU"] =datetime.date.today().__format__('Yeni Açılış' +" "+"("+ ('%d.%m.%Y'))+')'
        df_ana.loc[veri, "SAYISI"] = 1
ekle()
print("\n\nDF_ANA_KURUM NİHAİ HAL:\n", df_ana)
#------------------------------------------------------------------------------------------------------------------
def egitim_durumu(okul):
    if ("İlkokul" in okul) or ("I. Kademe" in okul):
        return "İLKOKUL"
    if ("Ortaokul" in okul) or ("II. Kademe" in okul):
        return "ORTAOKUL"
    if ("Lisesi" in okul) or ("Merkezi" in okul) or ("III. Kademe" in okul):
        return "ORTAÖĞRETİM"
df_ana["EĞİTİM_KADEMESİ"] = df_ana["KURUM_ADI"].apply(egitim_durumu)
#print("\n\nDF_ANA_KURUM EĞİTİM_KADEMESİ EKLENMİŞ NİHAİ HAL:\n", df_ana)
#---------------------------------------------------------------------------------------------------------------
df_ana=df_ana.set_index("KURUM_KODU")
df_ana=df_ana.sort_values(by=['İLÇE','KURUM_ADI'])

print(df_ana)
#------------------------------------------------------------------------------------------------------------
# Kaydet
def write_excel(filename, Sayfa, dataframe):
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer: 
        workBook = writer.book
        try:
            workBook.remove(workBook[Sayfa])
        except:
            print("Worksheet does not exist")
        finally:
            dataframe.to_excel(writer, sheet_name='Ana_Tablo')
            writer._save()
filename = ("Ana_Tablo.xlsx")
write_excel(filename, "Ana_Tablo" ,df_ana)

wb = load_workbook(filename)
sa = len(wb.sheetnames)     # sekme adları sayısı
wb.move_sheet("Ana_Tablo", offset = -(sa-1))
wb.save(filename)