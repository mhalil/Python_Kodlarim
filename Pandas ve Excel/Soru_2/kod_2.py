import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

dosya = pd.read_excel("ANA_TABLO.ods") 
df_ana= pd.DataFrame(dosya)

dosya = pd.read_excel("ÖĞRENCİ_SAYILARI.ods")
df_ogrenci = pd.DataFrame(dosya)  

df_ana["Kurum_Egitim"] = df_ana["KURUM_KODU"].astype(str) + "_" + df_ana["EĞİTİM_KADEMESİ"]   # df_ana["KURUM_KODU"].apply(str) şeklinde de kullanılabilir.
df_ana = df_ana.set_index("Kurum_Egitim")

df_ogrenci["Kurum_Egitim"] = df_ogrenci["KURUM_KODU"].apply(str) + "_" + df_ogrenci["EĞİTİM_KADEMESİ"]
df_ogrenci = df_ogrenci.set_index("Kurum_Egitim")

df_ana["ERKEK_ÖĞRENCİ"] = df_ogrenci.groupby("Kurum_Egitim").sum()["ERKEK_ÖĞRENCİ"]
df_ana["KIZ_ÖĞRENCİ"] = df_ogrenci.groupby("Kurum_Egitim").sum()["KIZ_ÖĞRENCİ"]
df_ana["TOPLAM_ÖĞRENCİ"] = df_ana["ERKEK_ÖĞRENCİ"] + df_ana["KIZ_ÖĞRENCİ"]
df_ana = df_ana.set_index("KURUM_KODU")

def write_excel(filename, Sayfa, dataframe):
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer: 
        workBook = writer.book
        try:
            workBook.remove(workBook[Sayfa])
        except:
            print("Worksheet does not exist")
        finally:
            dataframe.to_excel(writer, sheet_name=Sayfa, index=False)
            writer._save()

filename = "ANA_TABLO.xlsx"
write_excel(filename, "Sayfa1" ,df_ana)

# "Sayfa1" sekmesini, ilk sekmeye taşı
wb = load_workbook(filename)
sa = len(wb.sheetnames)     # sekme adları sayısı
wb.move_sheet("Sayfa1", offset = -(sa-1))
wb.save(filename)
